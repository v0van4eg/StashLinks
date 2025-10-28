# generators/base_generator.py
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class BaseGenerator:
    """Базовый класс для генерации XLSX документов"""

    def __init__(self, template_name=None):
        self.template_name = template_name
        # Используем путь из конфига, если шаблон указан
        if template_name:
            from config import Config
            self.template_path = Config.TEMPLATE_PATHS.get(template_name)
        else:
            self.template_path = None

    def load_template(self):
        """Загружает шаблон или создает новый документ"""
        if self.template_path and os.path.exists(self.template_path):
            print(f"Загружаем шаблон: {self.template_path}")
            wb = load_workbook(self.template_path)
            ws = wb.active
            start_row = self.get_start_row()
            return wb, ws, start_row
        else:
            print("Создаем новый документ (шаблон не найден)")
            return self.create_new_workbook()

    def create_new_workbook(self):
        """Создает новый документ с заголовками"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.get_worksheet_title()
        headers = self.get_headers()
        if headers:  # Добавляем заголовки только если они есть
            ws.append(headers)
            # Стили для заголовков
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        return wb, ws, self.get_start_row()

    def get_start_row(self):
        """Возвращает строку, с которой начинать заполнение данных"""
        return 2  # После заголовков

    def get_worksheet_title(self):
        """Возвращает название листа"""
        return "Images"

    def get_headers(self):
        """Возвращает заголовки столбцов"""
        raise NotImplementedError("Метод get_headers должен быть реализован в дочернем классе")

    def process_image_data(self, image_data):
        """Группирует изображения по артикулам"""
        articles = {}
        for item in image_data:
            article = item['article']
            if article not in articles:
                articles[article] = []
            articles[article].append(item['url'])
        return articles

    def generate(self, image_data, template_name):
        """Основной метод генерации документа"""
        try:
            print(f"Генерация XLSX для {len(image_data)} изображений, шаблон: {template_name}")

            wb, ws, start_row = self.load_template()
            articles = self.process_image_data(image_data)
            current_row = start_row

            print(f"Начинаем запись с строки {current_row}, артикулов: {len(articles)}")

            for article, urls in articles.items():
                row_data = self.generate_row_data(article, urls, template_name)
                self.write_row_data(ws, current_row, row_data)
                current_row += 1

            self.adjust_column_widths(ws)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            print("XLSX успешно сгенерирован")
            return buffer

        except Exception as e:
            print(f"Ошибка при генерации XLSX: {str(e)}")
            raise Exception(f"Error generating XLSX: {str(e)}")

    def generate_row_data(self, article, urls, template_name):
        """Генерирует данные для строки"""
        raise NotImplementedError("Метод generate_row_data должен быть реализован в дочернем классе")

    def write_row_data(self, ws, row_num, row_data):
        """Записывает данные в строку"""
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

    def adjust_column_widths(self, ws):
        """Настраивает ширину столбцов"""
        pass
