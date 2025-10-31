# generators/yandexmarket_generator.py
from .base_generator import BaseGenerator
from openpyxl.styles import Alignment
import io

class YandexmarketGenerator(BaseGenerator):
    def __init__(self, separator='comma'):
        super().__init__('В ячейку', separator)

    def get_worksheet_title(self):
        return "Images"

    def get_headers(self):
        # Только два столбца: Артикул и Ссылки
        return ["Артикул", "Ссылки на изображения"]

    def generate_row_data(self, article, urls, template_name):
        # Используем разделитель из настроек
        separator = self.get_separator()
        links_text = separator.join(urls) if urls else ""
        return [article, links_text]

    def adjust_column_widths(self, ws):
        # Настраиваем ширину в зависимости от разделителя
        if self.separator == 'newline':
            # Для переносов строк делаем колонку уже, но выше
            column_widths = {
                'A': 20,  # Артикул
                'B': 60,  # Ссылки (уже, так как вертикально)
            }

            # Устанавливаем перенос текста для ВСЕХ строк в колонке B
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=2)
                # Если ячейка не пустая, устанавливаем перенос текста
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            # Для запятых - широкая колонка
            column_widths = {
                'A': 20,  # Артикул
                'B': 100,  # Ссылки (широкая колонка для текста с запятыми)
            }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def generate(self, image_data, template_name):
        """Переопределяем метод generate для установки форматирования ВСЕХ строк"""
        try:
            print(f"Генерация XLSX для {len(image_data)} изображений, шаблон: {template_name}")

            wb, ws, start_row = self.load_template()
            articles = self.process_image_data(image_data)
            current_row = start_row

            print(f"Начинаем запись с строки {current_row}, артикулов: {len(articles)}")

            # Сортируем артикулы по алфавиту
            sorted_articles = sorted(articles.items(), key=lambda x: x[0])

            for article, urls in sorted_articles:
                row_data = self.generate_row_data(article, urls, template_name)
                self.write_row_data(ws, current_row, row_data)

                # Сразу устанавливаем форматирование для текущей строки
                if self.separator == 'newline' and len(urls) > 1:
                    cell = ws.cell(row=current_row, column=2)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                current_row += 1

            # Дополнительно вызываем adjust_column_widths для гарантии
            self.adjust_column_widths(ws)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            print("XLSX успешно сгенерирован")
            return buffer

        except Exception as e:
            print(f"Ошибка при генерации XLSX: {str(e)}")
            raise Exception(f"Error generating XLSX: {str(e)}")
